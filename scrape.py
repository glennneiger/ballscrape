""" BeautifulSoup is a module that allows you to extract data from html and
xml files. In this case, it'll be providing functions for navigating the
fucking mess that is a html file pulled from a website like this.

The platform attribute of the sys module returns a string representing the
running OS, which is useful because compatibility.

re module for regex

csv module for saving data
"""
from bs4 import BeautifulSoup
from sys import platform
import re
import csv

def has_xparam(tag):
    """ Function that returns true if the tag you pass it has both these attrs """
    return tag.has_attr('onclick') and tag.has_attr('xparam')

# Variable to store desired url
bball_url = "https://www.oddsportal.com/basketball/usa/nba/?fbclid=IwAR1LEc9UYgrWF8RYNsvyh7grWvExH1Ygw6iq9dl7jIIw_WGu5wR1dv9Q98k"

# This bit is pretty bloody slick. It checks your operating system and imports
# the relevant library to get the data from the webpage. Uses selenium for
# windows (not headless, will literally open a browser, load the webpage, then
# close the browser) and uses dryscrape for linux (headless, meaning it grabs
# the web data without displaying or opening anything to the user). The reason
# you have to use these libraries is that the betting odds website loads its
# odds using Javascript, meaning a simple http request won't cut it.
on_windows = False
on_linux = False

if platform == 'win32':
    on_windows = True
    from selenium import webdriver
    browser = webdriver.Firefox()
    browser.get(bball_url)
    response = browser.page_source
    browser.quit()
elif 'linux' in platform:
    on_linux = True
    import dryscrape
    session = dryscrape.Session()
    session.visit(bball_url)
    response = session.body()
else:
    raise RuntimeError('Use a normal fucking operating system you pleb')

# Give the raw content of the response to the BeautifulSoup module to make a
# 'BeautifulSoup' object. That BeautifulSoup object is what is going to do our
# navigation and return the information we want
soup = BeautifulSoup(response, 'html.parser')

# Makes a list of the children of the table, this is actually only two elements
# long. There's a colgroup which seems to define formatting for the columns,
# and then the table body which contains 'tr' elements, which are the rows in
# the table.
table = list(soup.find('table', class_='table-main').children)

# Pull out the table rows
table_rows = list(table[1].children)

# Empty variable to store the data we pull out to be saved later
matchups = []

for row in table_rows:
    # There are a bunch of rows in the table that are irrelevant so we need a
    # way to filter the actual match stats from the crap so we don't get
    # errors. The method here is to regex for a link that contains specific
    # text, text we know if going to be in a matchup name
    matchup = row.find(href=re.compile('basketball/usa/nba'))
    if matchup:
        odds = list(row.find_all(has_xparam))
        # There is an empty row or two that doesn't contain any odds, this line
        # just makes sure you skip over those
        if not odds:
            continue
        # Append matchup string, and odds to the variable from earlier
        matchups.append([matchup.text.split('-')[0].strip(), matchup.text.split('-')[1].strip(), float(odds[0].text.strip()), float(odds[1].text.strip())])
        
# Save the data to a CSV file
with open('game_odds.csv', 'w', newline='') as csvfile:
    writer = csv.writer(csvfile)
    for row in matchups:
        writer.writerow(row)

# Loading config from text file
workbook_name = ''
workbook_regex = re.compile(r'workbook_name=(.*)')
with open('config.txt', 'r') as config:
    for row in config:
        match = re.search(workbook_regex, row)
        if match:
            workbook_name = match.group(1)

if not workbook_name:
    raise RuntimeError('Could not find workbook name in config file')
workbook_name = workbook_name + '.xlsx'

# Only runs excel stuff on windows
if on_windows:
    from openpyxl import load_workbook, Workbook
    # Tries to find the workbook name you entered and creates a new one if it can't find it
    try:
        workbook = load_workbook(workbook_name)
    except FileNotFoundError as _:
        workbook = Workbook()
    if 'Scraped' not in workbook.sheetnames:
        sheet = workbook.create_sheet('Scraped')
    else:
        sheet = workbook["Scraped"]
    
    sheet_cells = []
    # Using iterables to access all our data and dump into spreadsheet, there might be a better way to do this but I'm bad at this
    iterable_matchups = iter(matchups)
    for row in sheet.iter_rows(max_col=len(matchups[0]), max_row=len(matchups)):
        values = iter(next(iterable_matchups))
        for cell in row:
            cell.value = next(values)

    workbook.save(workbook_name)